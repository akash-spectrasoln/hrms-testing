
from django.shortcuts import render
from django.http import HttpResponse
from.models import *
from django.shortcuts import render

# Create your views here.


# views.py


# views.py
# employee_app/views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.conf import settings
from django.utils import timezone
from django.forms import formset_factory
from django.db.models import Max,Min
from django.utils.decorators import method_decorator
from django.core.exceptions import ObjectDoesNotExist
from functools import wraps
from django.utils import timezone
import datetime
from django.db import IntegrityError
import time
#
# from django.shortcuts import render
from django.http import HttpResponse
from .models import Role, state, Country ,Employees,Department
from datetime import datetime,date
import time
TIMEOUT_DURATION = 60 * 60  # 60 minutes

def signin_required(fn):
    @wraps(fn)
    def wrapper(request, *args, **kwargs):
        # Check if the user is authenticated
        if not request.user.is_authenticated:
            return redirect("admin_login")
        
        # Check if the session has timed out due to inactivity
        current_time = time.time()
        last_activity = request.session.get('last_activity', current_time)

        if current_time - last_activity > TIMEOUT_DURATION:
            # If the session expired due to inactivity, log the user out
            logout(request)
            return redirect("admin_login")  # Redirect to sign-in page

        # Update the session's last activity timestamp
        request.session['last_activity'] = current_time

        # Proceed with the original view function
        return fn(request, *args, **kwargs)

    return wrapper


#
# index page view

def index(request):
    return render(request,'index.html')



def admin_index(request):
    return render(request,'admin_index.html')

# adding employees view


from django.contrib import messages

from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from .models import Employees, Salutation, Country, state, Department, Role
from .forms import EmployeeEditForm
from datetime import date, datetime

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import date

from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import date
from .forms import EmployeeEditForm
from .models import Role, Department, state, Country, Salutation, Employees
@signin_required
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, request.FILES, request=request)
        if form.is_valid():
            try:
                employee = form.save(commit=False) 
                # employee.data=employee.company_email  # Save the form but don't commit to the database yet
                employee.enc_home_city=employee.home_city
                employee.enc_personal_email=employee.personal_email 
                employee.enc_mobile_phone=employee.mobile_phone 
                employee.enc_home_phone=employee.home_phone  
                employee.enc_address=employee.address  
                employee.enc_pincode=employee.pincode  
                employee.enc_emergency_contact_phone=employee.emergency_contact_phone
                employee.enc_emergency_contact_name=employee.emergency_contact_name
                employee.enc_base_salary=employee.base_salary 
                employee.enc_pan_card=employee.pan_card  
                employee.enc_aadhaar=employee.aadhaar  
                employee.enc_bank_name=employee.bank_name   
                employee.enc_bank_branch=employee.bank_branch  
                employee.enc_bank_branch_address=employee.bank_branch_address   
                employee.enc_bank_account=employee.bank_account   
                employee.enc_ifsc_code=employee.ifsc_code 
                employee.enc_emergency_contact_relation=employee.emergency_contact_relation 
                employee.enc_incentive=employee.incentive 
                employee.enc_home_house=employee.home_house
                
                employee.enc_joining_bonus=employee.joining_bonus 

                employee.save()  # Now save the employee instance
                
                # Handle Admin Privileges
                is_admin = request.POST.get('is_admin') == 'on'  # Fetch and check the 'is_admin' checkbox
                print(is_admin)
                # Ensure the employee has an associated user object
                if employee.user:
                    user_obj = employee.user
                    user_obj.is_superuser = is_admin  # Set admin privileges as per checkbox state
                    user_obj.is_staff = is_admin  
                    user_obj.first_name = employee.first_name 
                    user_obj.last_name = employee.last_name 
                     # Typically set is_staff alongside is_superuser
                    user_obj.save()

                # Handle multiple resumes
                for resume_file in request.FILES.getlist('resumes'):
                    Resume.objects.create(employee=employee, file=resume_file)

                # Handle multiple certificates
                for certificate_file in request.FILES.getlist('certificates'):
                    Certificate.objects.create(employee=employee, file=certificate_file)

                messages.success(request, "Employee added successfully!")
                return redirect('add_employee')
            except Exception as e:
                messages.error(request, f"An error occurred while saving: {e}")
        else:
            # Print form errors in console for debugging
            
            print("Form Errors:", form.errors)
            messages.error(request, "Please correct the errors below.")
    else:
        form = EmployeeEditForm(request=request)

    # Fetch dropdown and default data
    context = {
        'form': form,
        'roles': Role.objects.all(),
        'departments': Department.objects.all(),
        'states': state.objects.all(),
        'countries': Country.objects.all(),
        'salutations': Salutation.objects.all(),
        'employees': Employees.objects.all(),
        'default_valid_from': date.today(),
        'default_valid_to': date(9999, 12, 31),
        'employee_types' : EmployeeType.objects.all().order_by('id')
    }
    return render(request, 'add_employee.html', context)



from django.http import JsonResponse
from .models import state  # Keep 'state' lowercase, as defined in your model

def get_states(request):
    country_id = request.GET.get('country_id')  # Get country ID from the request
    if country_id:
        states = state.objects.filter(country_id=country_id).values('id','name','code')  # Use 'state' lowercase
        print(states)  # Debugging line
        return JsonResponse(list(states), safe=False)  # Return states as JSON
    return JsonResponse({'error': 'Country not selected'}, status=400)







import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Employees

def export_employees_to_excel(request):
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'

    # Define the header row
    headers = [
        'Employee ID', 'Salutation', 'First Name', 'Middle Name', 'Last Name',
        'Date of Birth', 'Company Email', 'Personal Email', 'Mobile Phone', 'Office Phone', 
        'Home Phone', 'Valid From', 'Valid To', 'Country', 'State', 'Address', 
        'Home House', 'Home Post Office', 'Home City', 'Pincode', 'Role', 'Department', 
        'Manager', 'Employee Status', 'Emergency Contact Name', 'Emergency Contact Phone',
        'Emergency Contact Email', 'Emergency Contact Relation', 'Base Salary',
        'Created On', 'Modified On', 'Is Deleted', 'Floating Holidays Balance', 
        'Floating Holidays Used', 'Total Leaves', 'Used Leaves', 'Resignation Date',
        'Incentive', 'Joining Bonus', 'PAN Card', 'Aadhaar', 'Bank Name', 
        'Bank Branch', 'Bank Branch Address', 'Bank Account', 'IFSC Code'
    ]

    # Write the header row to the sheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Query all employees
    employees = Employees.objects.select_related('salutation', 'country', 'state', 'role', 'department', 'manager')

    # Write employee data to the sheet
    for row_num, employee in enumerate(employees, 2):
        # Convert datetime fields to naive datetime
        created_on = employee.created_on.replace(tzinfo=None) if employee.created_on else ''
        modified_on = employee.modified_on.replace(tzinfo=None) if employee.modified_on else ''

        row = [
            employee.employee_id,
            employee.salutation.sal_name if employee.salutation else '',
            employee.first_name,
            employee.middle_name,
            employee.last_name,
            employee.date_of_birth,
            employee.company_email,  # This is encrypted but accessed directly
            employee.enc_personal_email,  # Using encrypted property
            employee.enc_mobile_phone,    # Using encrypted property
            employee.office_phone,
            employee.enc_home_phone,      # Using encrypted property
            employee.valid_from,
            employee.valid_to,
            employee.country.country_name if employee.country else '',
            employee.state.name if employee.state else '',
            employee.enc_address,         # Using encrypted property
            employee.enc_home_house,      # Using encrypted property
            employee.home_post_office,    # Not encrypted in your model
            employee.enc_home_city,       # Using encrypted property
            employee.enc_pincode,         # Using encrypted property
            employee.role.role_name if employee.role else '',
            employee.department.dep_name if employee.department else '',
            f"{employee.manager.first_name} {employee.manager.last_name}" if employee.manager else '',
            employee.employee_status,
            employee.enc_emergency_contact_name,      # Using encrypted property
            employee.enc_emergency_contact_phone,     # Using encrypted property
            employee.emergency_contact_email,
            employee.enc_emergency_contact_relation,  # Using encrypted property
            employee.enc_base_salary,                 # Using encrypted property
            created_on,
            modified_on,
            employee.is_deleted,
            employee.floating_holidays_balance,
            employee.floating_holidays_used,
            employee.total_leaves,
            employee.used_leaves,
            employee.resignation_date,
            employee.enc_incentive,         # Using encrypted property
            employee.enc_joining_bonus,     # Using encrypted property
            employee.enc_pan_card,          # Using encrypted property
            employee.enc_aadhaar,           # Using encrypted property
            employee.enc_bank_name,         # Using encrypted property
            employee.enc_bank_branch,       # Using encrypted property
            employee.enc_bank_branch_address,  # Using encrypted property
            employee.enc_bank_account,      # Using encrypted property
            employee.enc_ifsc_code          # Using encrypted property
        ]

        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            # Handle None values and ensure proper data types
            if cell_value is None:
                sheet[f'{col_letter}{row_num}'] = ''
            elif isinstance(cell_value, str) and cell_value.replace('.', '').replace('-', '').isdigit():
                # Try to convert numeric strings to float for salary/incentive fields
                try:
                    sheet[f'{col_letter}{row_num}'] = float(cell_value)
                except ValueError:
                    sheet[f'{col_letter}{row_num}'] = cell_value
            else:
                sheet[f'{col_letter}{row_num}'] = cell_value

    # Set the response headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

from django.db.models import Q
from django.utils import timezone



from django.shortcuts import render
from django.db.models import Q
from django.utils import timezone
from .models import Employees

from django.shortcuts import render
from django.db.models import Q
from django.utils import timezone
from .models import Employees, Country  # Adjust as needed
from datetime import date
@signin_required
def list_employees(request):
    queryset = Employees.objects.filter(is_deleted=False).order_by('-modified_on')
    country_list = Country.objects.all().order_by('country_name')

    # Get filter parameters
    emp_id = request.GET.get('employee_id', '').strip()
    name = request.GET.get('name', '').strip()
    status = request.GET.get('employee_status', 'employed')
    country_id = request.GET.get('country_id', '').strip()
    print(country_id)
    # Default to India's id if no country selected
    # Default to India's id if no country is selected
    if not country_id:
        india = Country.objects.filter(country_name__iexact='India').first()
        country_id = str(india.id) if india else ''
    elif not country_id.isdigit():
        country_id = ''

    # Status filter
    if status != 'all':
        queryset = queryset.filter(employee_status=status)
    # Employee ID filter
    if emp_id:
        queryset = queryset.filter(employee_id__icontains=emp_id)
    # Name filter
    if name:
        queryset = queryset.filter(
            Q(first_name__icontains=name) |
            Q(last_name__icontains=name)
        )
    # Country filter
    if country_id:
        queryset = queryset.filter(country__id=country_id)

    today = timezone.now().date()
    employee_list = []

    for employee in queryset.select_related('manager', 'department', 'role', 'country'):
        if not employee.pk:
            continue

        manager_display = (
            f"{employee.manager.employee_id} ({employee.manager.first_name})"
            if employee.manager else "None"
        )

        resignation_tooltip = ""
        if employee.employee_status == 'resigned' and employee.resignation_date:
            formatted_date = employee.resignation_date.strftime('%b %d, %Y')
            days_ago = (today - employee.resignation_date).days
            resignation_tooltip = f"Resigned on {formatted_date} ({days_ago} days ago)"

        employee_list.append({
            'pk': employee.pk,
            'employee_id': employee.employee_id,
            'first_name': employee.first_name,
            'last_name': employee.last_name,
            'company_email': employee.company_email,
            'mobile_phone': employee.enc_mobile_phone,
            'department': employee.department.dep_name if employee.department else "",
            'designation': employee.role.role_name if employee.role else "",
            'country': employee.country.country_name if employee.country else "",
            'employee_status': employee.employee_status,
            'manager_display': manager_display,
            'resignation_tooltip': resignation_tooltip,
            'resignation_date': employee.resignation_date,
            'today': date.today(),
        })

    context = {
        'employee_list': employee_list,
        'country_list': country_list,
        'current_filters': {
            'emp_id': emp_id,
            'name': name,
            'status': status,
            'country_id': country_id,
        }
    }

    return render(request, 'list_employees.html', context)

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from datetime import date
from .models import LeaveRequest, Employees  # Adjust this import if your app name/model name differs
@signin_required
def delete_leave_request(request, leave_id):
    leave = get_object_or_404(LeaveRequest, id=leave_id)
    employee = leave.employee_user
    leave_days = leave.leave_days  # Ensure this exists in your model
    try:
        current_user = Employees.objects.get(user=request.user)
    except Employees.DoesNotExist:
        messages.error(request, "You are not registered as an employee.")
        return redirect("leave_request_display")

    try:
        emp_obj = Employees.objects.get(company_email=employee.email)
    except Employees.DoesNotExist:
        emp_obj = None

    if emp_obj:
        if leave.leave_type == "Casual Leave":
            emp_obj.used_leaves = max(emp_obj.used_leaves - leave_days, 0)
            emp_obj.save()
        elif leave.leave_type == "Floating Leave":
            emp_obj.floating_holidays_used = max(emp_obj.floating_holidays_used - leave_days, 0)
            emp_obj.save()
        # ...handle other leave types as needed

    # Soft delete instead of hard delete
    leave.status='Deleted'
    leave.approved_by=current_user.user
    leave.save()
    messages.success(request, "Leave request deleted successfully.")

    return redirect("leave_request_display")







from django.views.generic import UpdateView
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from .models import Employees, Country, state
from .forms import EmployeeEditForm




from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.core.files.storage import default_storage
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError

# Import models
from .models import (
    Employees, Salutation, Role, Department,
    Country, state
)
from .forms import EmployeeEditForm

import logging
import traceback

# Configure logging
logger = logging.getLogger(__name__)


from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.db import transaction
from django.core.files.storage import default_storage
from django.utils import timezone
from django.core.exceptions import ValidationError
import logging
import traceback
from .models import Employees, Salutation, Country, state, Department, Role
from .forms import EmployeeEditForm
from django.utils.decorators import method_decorator




logger = logging.getLogger(__name__)
@method_decorator(signin_required, name='dispatch')
class EmployeeUpdateView(UpdateView):
    model = Employees
    form_class = EmployeeEditForm
    template_name = 'employeeupdate.html'
    success_url = reverse_lazy('employee_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            context['employee_type'] = EmployeeType.objects.all()
            context['salutations'] = Salutation.objects.all()
            context['roles'] = Role.objects.all()
            context['departments'] = Department.objects.all()
            context['managers'] = Employees.objects.filter(employees_managed__isnull=False).distinct()
            context['employees'] = Employees.objects.all()
            employee = self.object
            context['countries'] = Country.objects.all()
            context['selected_country'] = employee.country if employee.country else None
            context['states'] = state.objects.filter(country=employee.country) if employee.country else []
            context['selected_state'] = employee.state if employee.state else None
            context['home_post_office'] = employee.home_post_office
            context['incentive'] = employee.enc_incentive
            context['joining_bonus'] = employee.enc_joining_bonus
            
            return context
        except Exception as e:
            logger.error(f"Error in get_context_data: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(self.request, f"An error occurred while preparing the form: {str(e)}")
            raise

    @transaction.atomic
    def form_valid(self, form):
        try:
            employee = form.save(commit=False)
            employee.enc_home_city=employee.home_city
            employee.enc_personal_email=employee.personal_email 
            employee.enc_mobile_phone=employee.mobile_phone 
            employee.enc_home_phone=employee.home_phone  
            employee.enc_address=employee.address  
            employee.enc_pincode=employee.pincode  
            employee.enc_emergency_contact_phone=employee.emergency_contact_phone
            employee.enc_emergency_contact_name=employee.emergency_contact_name
            employee.enc_base_salary=employee.base_salary 
            employee.enc_pan_card=employee.pan_card  
            employee.enc_aadhaar=employee.aadhaar  
            employee.enc_bank_name=employee.bank_name   
            employee.enc_bank_branch=employee.bank_branch  
            employee.enc_bank_branch_address=employee.bank_branch_address   
            employee.enc_bank_account=employee.bank_account   
            employee.enc_ifsc_code=employee.ifsc_code 
            employee.enc_emergency_contact_relation=employee.emergency_contact_relation 
            employee.enc_incentive=employee.incentive 
            employee.enc_home_house=employee.home_house
            
            employee.enc_joining_bonus=employee.joining_bonus 

            employee.save() 
            current_status = form.cleaned_data.get('employee_status')
            new_employee_type =  form.cleaned_data.get("employee_type")
            if current_status == 'employed':
                employee.resignation_date = None
            elif current_status == 'resigned' and not employee.resignation_date:
                employee.resignation_date = timezone.now().date()

            if new_employee_type and new_employee_type.code != employee.employee_type.code:
                unique_number = employee.employee_id[-4:]
                employee.employee_id = f"{new_employee_type}{unique_number}"

            if 'delete_resumes' in self.request.POST:
                for resume_id in self.request.POST.getlist('delete_resumes'):
                    try:
                        resume = Resume.objects.get(id=resume_id, employee=employee)
                        resume.file.delete()
                        resume.delete()
                    except Resume.DoesNotExist:
                        continue

            if 'delete_certificates' in self.request.POST:
                for certificate_id in self.request.POST.getlist('delete_certificates'):
                    try:
                        certificate = Certificate.objects.get(id=certificate_id, employee=employee)
                        certificate.file.delete()
                        certificate.delete()
                    except Certificate.DoesNotExist:
                        continue

            if self.request.FILES.getlist('resumes'):
                for resume_file in self.request.FILES.getlist('resumes'):
                    Resume.objects.create(employee=employee, file=resume_file)

            if self.request.FILES.getlist('certificates'):
                for certificate_file in self.request.FILES.getlist('certificates'):
                    Certificate.objects.create(employee=employee, file=certificate_file)

            employee.save()

            # -------- ADMIN PRIVILEGE HANDLING --------
            is_admin = self.request.POST.get('is_admin') == 'on'
            print("admin or not",is_admin)
            if hasattr(employee, 'user') and employee.user is not None:
                user_obj = employee.user
                user_obj.is_superuser = is_admin
                user_obj.is_staff = is_admin
                user_obj.first_name=employee.first_name
                user_obj.last_name=employee.last_name
                user_obj.save()
            # ------------------------------------------

            messages.success(self.request, "✅ Employee details updated successfully.")
            return super().form_valid(form)

        except ValidationError as ve:
            messages.error(self.request, f"Validation Error: {str(ve)}")
            return self.form_invalid(form)

        except Exception as e:
            logger.error(f"Unexpected error in form_valid: {str(e)}")
            logger.error(traceback.format_exc())
            transaction.set_rollback(True)
            messages.error(self.request, f"An unexpected error occurred: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        logger.error("Form Validation Failed")
        logger.error("Form Errors: %s", form.errors)
        logger.error("POST Data: %s", self.request.POST)

        if form.errors:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(self.request, f"Error in {field}: {error}")

        messages.error(self.request, "There was an error updating the employee details. Please check the form.")
        return super().form_invalid(form)
    
from django.http import JsonResponse
from django.shortcuts import render
from django.views import View
import pandas as pd
from .models import Employees, Country, state, Role, Department, Salutation

from django.views import View
from django.shortcuts import render
from django.http import JsonResponse
import pandas as pd
from .models import Employees, Country, state, Role, Department, Salutation
from django.utils.dateparse import parse_date
from datetime import datetime
import math
from django.views import View
from django.shortcuts import render
from django.http import JsonResponse
from django.db import transaction
from django.utils.dateparse import parse_date
from datetime import datetime
import pandas as pd
import math

from .models import Employees, Country, state, Role, Department, Salutation


from django.views import View
from django.shortcuts import render
from django.http import JsonResponse
from django.db import transaction
from django.utils.dateparse import parse_date
from datetime import datetime
import pandas as pd
import math

from .models import Employees, Country, state, Department, Salutation, Role

class EmployeeExcelCreateView(View):
    template_name = 'employee_excel_create.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        import traceback

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return JsonResponse({'error': "Please upload an Excel file."}, status=400)

        created_employees = []
        failed_rows = []

        required_fields = [
            'Employee Type',  'Salutation ID', 'First Name', 'Last Name',
            'Company Email', 'Personal Email', 'Mobile Phone',
            'Valid From', 'Valid To', 'Country Code', 'State Code',
            'Home City','State Code','Date Of Birth',
            'Department ID',  'Base Salary',
            'Bank Name', 'Bank Branch', 'Bank Branch Address', 'Bank Account', 'IFSC Code'
        ]

        try:
            df = pd.read_excel(excel_file)
            df = df.where(pd.notnull(df), None)  # Now, all NaNs are None
            df.columns = [c.strip() for c in df.columns]
            print("[DEBUG] Excel columns:", list(df.columns))

            for index, row in df.iterrows():
                try:
                    
                    excel_row_num = int(index) + 2
                    print(f"[DEBUG] Processing row {excel_row_num} ")
                    
                    # Check for missing required fields
                    missing_fields = [
                        field for field in required_fields
                        if field not in row or pd.isnull(row[field]) or row[field] == ''
                    ]
                    if missing_fields:
                        print(f"[DEBUG] Row {excel_row_num}: Missing fields: {missing_fields}")
                        failed_rows.append({
                            "row": excel_row_num,
                            
                            "error": "Missing required fields: " + ", ".join(missing_fields)
                        })
                        continue

                    # Type validation
                    type_errors = []
                    for date_field in ['Valid From', 'Valid To', 'Date of Birth', 'Resignation Date']:
                        if date_field in row and not pd.isnull(row[date_field]) and \
                            not isinstance(row[date_field], (pd.Timestamp, datetime)) and not parse_date(str(row[date_field])):
                            type_errors.append(f'"{date_field}": "{row[date_field]}"')
                    try:
                        float(row['Base Salary'])
                    except Exception as te:
                        print(f"[DEBUG] Row {excel_row_num}: Invalid Base Salary: {row['Base Salary']}")
                        type_errors.append(f'"Base Salary": "{row["Base Salary"]}"')
                    if 'Incentive' in row and not pd.isnull(row['Incentive']):
                        try:
                            float(row['Incentive'])
                        except Exception as te:
                            print(f"[DEBUG] Row {excel_row_num}: Invalid Incentive: {row['Incentive']}")
                            type_errors.append(f'"Incentive": "{row["Incentive"]}"')

                    if type_errors:
                        print(f"[DEBUG] Row {excel_row_num}: Type errors: {type_errors}")
                        failed_rows.append({
                            "row": excel_row_num,
                            
                            "error": "Invalid value(s): " + ", ".join(type_errors)
                        })
                        continue

                    # FK LOOKUPS
                    try:
                        print(f"[DEBUG] Row {excel_row_num}: Fetching FKs")
                        country = Country.objects.get(code=row['Country Code'])
                        print(f"   Country: {country}")
                        state_obj = state.objects.get(code=(row['State Code']),country_id=country.id)
                        print(f"   State: {state_obj}")
                        department = Department.objects.get(id=int(row['Department ID']))
                        print(f"   Department: {department}")
                        salutation = Salutation.objects.get(id=int(row['Salutation ID']))
                        print(f"   Salutation: {salutation}")
                        employee_type_obj=EmployeeType.objects.get(id=int(row['Employee Type']))
                        role_obj = None
                        role_obj = Role.objects.get(role_id=int(row['Role ID']))
                        manager=None
                        
                        print(row.get('Manager', ''))
                        manager = Employees.objects.get(employee_id=row['Manager']) if row.get('Manager') else None   # remove whitespace
                    except Exception as e:
                        debug_info = traceback.format_exc()
                        print(f"[DEBUG] Row {excel_row_num}: Foreign key error:\n{debug_info}")
                        failed_rows.append({
                            "row": excel_row_num,
                            
                            "error": f"Foreign key lookup error: {e}"
                        })
                        continue

                    # Helpers
                    def parse(val): return None if pd.isnull(val) or val == '' else val
                    def parse_decimal(val):
                        try: return float(val)
                        except: return 0.0
                    def parse_date_field(val):
                        if pd.isnull(val) or not val: return None
                        if isinstance(val, (pd.Timestamp, datetime)): return val.date()
                        return parse_date(str(val))

                    try:
                        print(f"[DEBUG] Row {excel_row_num}: Creating Employee object")
                        with transaction.atomic():
                            # Create employee with non-encrypted fields first
                            employee = Employees(
                                employee_type=employee_type_obj,
                                employee_id=generate_employee_id(employee_type_obj.code),
                                old_employee_id=parse(row["Old Employee Id"]) if 'Old Employee Id' in row else '',
                                salutation=salutation,
                                first_name=parse(row['First Name']),
                                middle_name=parse(row['Middle Name']) if 'Middle Name' in row else '',
                                last_name=parse(row['Last Name']),
                                company_email=parse(row['Company Email']),  # This is stored encrypted but no property exists
                                office_phone=str(parse(row['Office Phone'])) if 'Office Phone' in row else '',
                                valid_from=parse_date_field(row['Valid From']),
                                valid_to=parse_date_field(row['Valid To']),
                                country=country,
                                state=state_obj,
                                home_post_office=parse(row['Home Post Office']),
                                department=department,
                                role=role_obj,
                                manager=manager,
                                date_of_birth=parse_date_field(row['Date Of Birth']) if 'Date Of Birth' in row else None,
                                resignation_date=parse_date_field(row['Resignation Date']) if 'Resignation Date' in row else None,
                                emergency_contact_email=parse(row['Emergency Contact Email']) if 'Emergency Contact Email' in row else None,
                            )
                            
                            # Set encrypted fields using the properties
                            # Personal information
                            employee.enc_personal_email = parse(row['Personal Email'])
                            employee.enc_mobile_phone = str(parse(row['Mobile Phone']))
                            employee.enc_home_phone = str(parse(row['Home Phone'])) if 'Home Phone' in row else ''
                            
                            # Address information
                            employee.enc_address = parse(row['Address']) if 'Address' in row else ''
                            employee.enc_home_house = parse(row['Home House'])
                            employee.enc_home_city = parse(row['Home City'])
                            employee.enc_pincode = str(parse(row['Pincode'])) if 'Pincode' in row else ''
                            
                            # Emergency contact
                            employee.enc_emergency_contact_name = parse(row['Emergency Contact Name']) if 'Emergency Contact Name' in row else None
                            employee.enc_emergency_contact_phone = str(parse(row['Emergency Contact Phone'])) if 'Emergency Contact Phone' in row else ''
                            employee.enc_emergency_contact_relation = parse(row['Emergency Contact Relation']) if 'Emergency Contact Relation' in row else None
                            
                            # Financial information
                            employee.enc_base_salary = str(parse_decimal(row['Base Salary']))
                            employee.enc_joining_bonus = str(parse_decimal(row['Joining Bonus'])) if 'Joining Bonus' in row else '0.0'
                            employee.enc_incentive = str(parse_decimal(row['Incentive'])) if 'Incentive' in row else '0.0'
                            
                            # Government IDs
                            employee.enc_pan_card = parse(row['PAN Card']) if 'PAN Card' in row else ''
                            employee.enc_aadhaar = parse(row['Aadhar']) if 'Aadhar' in row else ''
                            
                            # Bank information
                            employee.enc_bank_name = parse(row['Bank Name']) if 'Bank Name' in row else ''
                            employee.enc_bank_branch = parse(row['Bank Branch']) if 'Bank Branch' in row else ''
                            employee.enc_bank_branch_address = parse(row['Bank Branch Address']) if 'Bank Branch Address' in row else ''
                            employee.enc_bank_account = parse(row['Bank Account']) if 'Bank Account' in row else ''
                            employee.enc_ifsc_code = parse(row['IFSC Code']) if 'IFSC Code' in row else ''
                            
                            employee.save()

                            if employee.user:
                                user_obj = employee.user
                                
                                user_obj.first_name = employee.first_name 
                                user_obj.last_name = employee.last_name  # Typically set is_staff alongside is_superuser
                                user_obj.save()
                        print(f"[DEBUG] Row {excel_row_num}: Employee created: {employee}")

                        created_employees.append({
                            'row': excel_row_num,
                            'employee_id': employee.employee_id,
                            'first_name': employee.first_name,
                            'last_name': employee.last_name,
                            'department': employee.department.dep_name if employee.department else '',
                            'role': str(employee.role) if employee.role else '',
                            'company_email': employee.company_email,
                            'mobile_phone': employee.enc_mobile_phone  # Use encrypted property for display
                        })

                    except Exception as e:
                        debug_info = traceback.format_exc()
                        print(f"[DEBUG] Row {excel_row_num}: Exception in model save:\n{debug_info}")
                        failed_rows.append({
                            'row': excel_row_num,
                            
                            'error': f"Error on create/save: {e}"
                        })

                except Exception as row_e:
                    debug_info = traceback.format_exc()
                    print(f"[DEBUG] Fatal error for row {excel_row_num}:\n{debug_info}")
                    failed_rows.append({
                        'row': excel_row_num,
                        
                        'error': f"Fatal row error: {row_e}"
                    })

            return JsonResponse({
                'created_employees': created_employees,
                'failed_rows': failed_rows,
            })

        except Exception as e:
            import traceback
            print("[DEBUG] Top-level exception in POST handler:")

import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from .models import Holiday, FloatingHoliday, StateHoliday, Country, state  # import your models

class HolidayExcelCreateView(View):
    template_name = 'holiday_excel_create.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect(request.path)

        try:
            df = pd.read_excel(excel_file)
            df = df.where(pd.notnull(df), None)  # Replace NaNs with None
            df.columns = [c.strip() for c in df.columns]  # Clean column names

            success_rows = []
            failed_rows = []

            for index, row in df.iterrows():
                try:
                    name = row['name']
                    leave_type = row['leave_type'].strip().lower()
                    date = pd.to_datetime(row['date']).date()
                    year = date.year
                    day = date.strftime("%A")

                    country = Country.objects.get(code=row['country_code'])

                    if leave_type == 'country':
                        if not Holiday.objects.filter(date=date, country=country).exists():
                            Holiday.objects.create(
                                name=name,
                                date=date,
                                day=day,
                                year=year,
                                country=country
                            )
                            success_rows.append({'name':name,'type':'Counrty'})
                        else:
                            failed_rows.append({'name':name})

                    elif leave_type == 'floating':
                        if not FloatingHoliday.objects.filter(date=date, country=country).exists():
                
                            FloatingHoliday.objects.create(
                                name=name,
                                date=date,
                                year=year,
                                country=country
                            )
                            success_rows.append({'name':name,'type':'Floating'})
                        else:
                            failed_rows.append({'name':name})

                    elif leave_type == 'state':
                        holiday_state= state.objects.get(code=row['State Code']) # holiday_state - state to which holiday is added
                        if not StateHoliday.objects.filter(date=date, country=country, state=holiday_state).exists():
                            StateHoliday.objects.create(
                                name=name,
                                date=date,
                                year=year,
                                country=country,
                                state=holiday_state
                            )
                            success_rows.append({'name':name,'type':'State'})
                        else:
                            failed_rows.append({'name':name})

                except Exception as e:
                    print(f"failed: {e}")
                    traceback.print_exc()

            messages.success(request, f"✅ {len(success_rows)} holidays uploaded successfully!")
            if failed_rows:
                messages.warning(request, f"⚠️ Failed to upload rows: {failed_rows}")
            return render(request, self.template_name, {
                                "success_names": success_rows,
                                "failed_rows": failed_rows
                            })

        except Exception as e:
            messages.error(request, f"❌ Error reading file: {e}")
            return redirect(request.path)


from django.contrib import messages
from django.views import View
from .models import Employees

class EmployeeExcelUpdateView(View):
    template_name = 'employee_excel_update.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('employee_excel_update')

        try:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(excel_file)

            # Validate and update each employee
            for index, row in df.iterrows():
                try:
                    employee_id = row['employee_id']
                    employee = Employees.objects.get(employee_id=employee_id)

                    # Update fields
                    employee.first_name = row.get('first_name', employee.first_name)
                    employee.middle_name = row.get('middle_name', employee.middle_name)
                    employee.last_name = row.get('last_name', employee.last_name)
                    employee.company_email = row.get('company_email', employee.company_email)
                    employee.personal_email = row.get('personal_email', employee.personal_email)
                    employee.mobile_phone = row.get('mobile_phone', employee.mobile_phone)
                    employee.office_phone = row.get('office_phone', employee.office_phone)
                    employee.home_phone = row.get('home_phone', employee.home_phone)
                    employee.home_city = row.get('home_city', employee.home_city)
                    employee.home_post_office = row.get('home_post_office', employee.home_post_office)
                    employee.home_house = row.get('home_house', employee.home_house)
                    employee.pincode = row.get('pincode', employee.pincode)
                    employee.department_id = row.get('department_id', employee.department_id)
                    employee.designation = row.get('designation', employee.designation)
                    employee.manager_id = row.get('manager_id', employee.manager_id)
                    employee.employee_status = row.get('employee_status', employee.employee_status)
                    employee.emergency_contact_name = row.get('emergency_contact_name', employee.emergency_contact_name)
                    employee.emergency_contact_phone = row.get('emergency_contact_phone', employee.emergency_contact_phone)
                    employee.emergency_contact_relation = row.get('emergency_contact_relation', employee.emergency_contact_relation)
                    employee.base_salary = row.get('base_salary', employee.base_salary)
                    employee.employee_type = row.get('employee_type', employee.employee_type)
                    employee.resignation_date = row.get('resignation_date', employee.resignation_date)
                    employee.valid_from = row.get('valid_from', employee.valid_from)
                    employee.valid_to = row.get('valid_to', employee.valid_to)
                    employee.country_id = row.get('country_id', employee.country_id)
                    employee.state_id = row.get('state_id', employee.state_id)

                    employee.save()
                except Employees.DoesNotExist:
                    messages.warning(request, f"Employee with ID {employee_id} does not exist.")
                except Exception as e:
                    messages.error(request, f"Error updating employee {employee_id}: {str(e)}")

            messages.success(request, "Employee records updated successfully.")
        except Exception as e:
            messages.error(request, f"Failed to process the Excel file: {str(e)}")

        return redirect('employee_excel_update')

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Resume, Certificate  # Import your models

@csrf_exempt
def delete_file(request):
    if request.method == 'POST':
        file_id = request.POST.get('file_id')
        file_type = request.POST.get('file_type')

        try:
            if file_type == 'resume':
                file_instance = Resume.objects.get(id=file_id)
            elif file_type == 'certificate':
                file_instance = Certificate.objects.get(id=file_id)
            else:
                return JsonResponse({'success': False})

            file_instance.delete()
            return JsonResponse({'success': True})
        except (Resume.DoesNotExist, Certificate.DoesNotExist):
            return JsonResponse({'success': False})

    return JsonResponse({'success': False})


#
# from django.http import JsonResponse
# from .models import state
#
# def get_states(request):
#     country_id = request.GET.get('country_id')
#     states = state.objects.filter(country_id=country_id).values('id', 'name')
#     return JsonResponse({'states': list(states)})




# from django.views.generic import DeleteView
# from .models import Employees
#
# class EmployeeDeleteView(DeleteView):
#     model = Employees
#     template_name = 'delete_employee.html'
#     context_object_name = 'data'
#     success_url = reverse_lazy('employee_list')  # Redirect after delete
#
#     #soft delete logic mentioned below
#
#     def delete(self, request, *args, **kwargs):
#         """Mark the employee as deleted instead of deleting from the database."""
#         self.object = self.get_object()
#         self.object.delete()  # Call the soft delete method in the model
#
#
#         return redirect(self.success_url)





from django.urls import reverse_lazy
from django.shortcuts import get_object_or_404, redirect
from django.views.generic import DeleteView
from django.contrib import messages
from .models import Employees
@method_decorator(signin_required, name='dispatch')
class EmployeeDeleteView(DeleteView):
    model = Employees
    template_name = 'delete_employee.html'
    context_object_name = 'data'
    success_url = reverse_lazy('employee_list')  # Redirect URL after delete

    def post(self, request, *args, **kwargs):
        """Handle POST request for soft delete and show a success message."""
        employee = get_object_or_404(Employees, pk=self.kwargs['pk'])  # Fetch employee
        emp_name = f"{employee.first_name} {employee.last_name}"  # Get full name

        if not employee.is_deleted:
            employee.is_deleted = True  # Soft delete
            employee.save()
            messages.success(request, f"✅ Employee {emp_name} has been deleted.")  # Success message
        else:
            messages.info(request, "ℹ️ This employee is already deleted.")  # Already deleted message

        return redirect(self.success_url)  # Redirect to employee list
#


from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .models import Employees

def restore_employee(request, pk):
    """Restore a soft-deleted employee."""
    employee = get_object_or_404(Employees, pk=pk)
    if employee.is_deleted:
        employee.is_deleted = False
        employee.save()
        messages.success(request, f"✅ Employee {employee.first_name} {employee.last_name} has been restored.")
    else:
        messages.info(request, "This employee is already active.")
    return redirect('deleted_employees')

from django.views.generic.list import ListView
from .models import Employees
@method_decorator(signin_required, name='dispatch')
class DeletedEmployeeListView(ListView):
    model = Employees
    template_name = "deleted_employees_display.html"
    context_object_name = "employees"

    def get_queryset(self):
        return Employees.objects.filter(is_deleted=True)

# Admin Dashboard View (after successful login)
from django.contrib.auth.decorators import login_required

@login_required
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')  # Render the admin dashboard page


#


from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from datetime import timedelta
from .models import LeaveRequest, Employees
from employee_app.models import Holiday, FloatingHoliday

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from datetime import timedelta
@signin_required
def accept_leave_request(request, leave_request_id):
    if request.method == "POST":
        leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
        employee = leave_request.employee_master

        # Fetch policy for floating holidays
        floating_entitlement = getattr(employee, 'floating_holidays_balance', 2)  # Or from policy table if relevant

        # Get all holiday/floating dates for efficiency
        holiday_dates = set(Holiday.objects.values_list('date', flat=True))
        floating_dates = set(FloatingHoliday.objects.values_list('date', flat=True))

        casual_days = 0
        floating_days = 0

        floating_h_used = employee.floating_holidays_used  # initial used

        current_date = leave_request.start_date
        while current_date <= leave_request.end_date:
            if current_date.weekday() in [5, 6]:  # weekend
                current_date += timedelta(days=1)
                continue
            if current_date in holiday_dates:
                current_date += timedelta(days=1)
                continue
            if current_date in floating_dates and floating_h_used + floating_days < floating_entitlement:
                floating_days += 1
                current_date += timedelta(days=1)
                continue
            casual_days += 1
            current_date += timedelta(days=1)

        # Check employee has enough balance
        if employee.used_leaves + casual_days > employee.total_leaves:
            messages.error(request, "Employee cannot exceed the allowed casual leave days.")
            return redirect('leave_request_display')
        if employee.floating_holidays_used + floating_days > employee.floating_holidays_balance:
            messages.error(request, "Employee cannot exceed the allowed floating holidays.")
            return redirect('leave_request_display')

        # Update leave usage
        employee.used_leaves += casual_days
        employee.floating_holidays_used += floating_days
        employee.save()

        # Update leave request (status, approver, leave_days)
        leave_request.status = "Approved"
        leave_request.approved_by = request.user
        leave_request.leave_days = casual_days + floating_days  # or just "leave_days" if consistent
        leave_request.save()

        # Email to Employee
        employee_email = employee.user.email
        employee_subject = "Leave Request Approved"
        employee_message = (
            f"Dear {employee.first_name},\n\n"
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been approved.\n"
            f"Status: Approved\n"
            f"Casual Leave Days: {casual_days}\n"
            f"Floating Holiday Days: {floating_days}\n"
            f"Total Leave Days: {casual_days + floating_days}\n"
            f"Enjoy your leave!\n\n"
            f"Best Regards,\nYour Leave Management System"
        )
        
        # Send email to employee
        send_mail(employee_subject, employee_message, settings.DEFAULT_FROM_EMAIL, [employee_email])
        
        # Email to PM (if PM email exists)
        if employee.pm_email:
            pm_subject = f"Leave Request Approved - {employee.first_name} {employee.last_name}"
            pm_message = (
                f"Dear Project Manager,\n\n"
                f"{employee.first_name} {employee.last_name}'s leave request has been approved.\n\n"
                f"Employee ID: {employee.employee_id}\n"
                f"Leave Period: {leave_request.start_date} to {leave_request.end_date}\n"
                f"Casual Leave Days: {casual_days}\n"
                f"Floating Holiday Days: {floating_days}\n"
                f"Total Leave Days: {casual_days + floating_days}\n"
                f"Reason: {leave_request.reason}\n\n"
                f"Best Regards,\nYour Leave Management System"
            )
            
            # Send email to PM
            send_mail(pm_subject, pm_message, settings.DEFAULT_FROM_EMAIL, [employee.pm_email])
        
        messages.success(request, "The leave request has been approved.")

        return redirect('leave_request_display')

    return HttpResponse("Invalid Request", status=400)

@signin_required
def reject_leave_request(request, leave_request_id):
    if request.method == "POST":
        leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
        leave_request.status = "Rejected"
        leave_request.approved_by = request.user  # Set who rejected
        leave_request.save()

        # Get employee object
        employee = leave_request.employee_master
        employee_email = employee.user.email

        # Email to Employee
        employee_subject = "Leave Request Rejected"
        employee_message = (
            f"Dear {employee.first_name},\n\n"
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been rejected.\n"
            f"Status: Rejected\n"
            f"If you have any questions or would like to discuss this decision, please contact HR.\n\n"
            f"Best Regards,\nYour Leave Management System"
        )

        # Send email to employee
        send_mail(employee_subject, employee_message, settings.DEFAULT_FROM_EMAIL, [employee_email])

        # Email to PM (if PM email exists)
        if employee.pm_email:
            pm_subject = f"Leave Request Rejected - {employee.first_name} {employee.last_name}"
            pm_message = (
                f"Dear Project Manager,\n\n"
                f"{employee.first_name} {employee.last_name}'s leave request has been rejected.\n\n"
                f"Employee ID: {employee.employee_id}\n"
                f"Leave Period: {leave_request.start_date} to {leave_request.end_date}\n"
                f"Reason for Leave: {leave_request.reason}\n"
                f"Status: Rejected\n\n"
                f"The employee has been notified of this decision.\n\n"
                f"Best Regards,\nYour Leave Management System"
            )
            
            # Send email to PM
            send_mail(pm_subject, pm_message, settings.DEFAULT_FROM_EMAIL, [employee.pm_email])

        messages.success(request, "The leave request has been rejected.")

        return redirect('leave_request_display')  # Redirect back to the display page

    return HttpResponse("Invalid Request", status=400)



from django.shortcuts import redirect
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required

def admin_logout(request):
    logout(request)  # Logs out the admin user
    return redirect('admin_login')  # Redirect to the admin login page after logout




  # Ensure this form is correctly defined in your forms.py
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from datetime import datetime
from .models import Holiday, FloatingHoliday , StateHoliday
from .forms import Holiday_Form



from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
@signin_required
def add_holidays(request):
    # Figure out initial/current filter values:
    # 1) POST (user submitting form, which may error)
    # 2) GET (came here via redirect after successful save)
    # 3) Default (current year, blank country/type)
    current_year = datetime.now().year
    get = request.GET
    post = request.POST

    # Handle persistence of filters:
    filter_country = post.get('country') or get.get('country') or ''
    filter_leave_type = post.get('leave_type') or get.get('leave_type') or ''
    filter_year = post.get('year') or get.get('year') or str(current_year)

    if request.method == "POST":
        form = Holiday_Form(request.POST)
        if form.is_valid():
            leave_type = form.cleaned_data['leave_type']
            name = form.cleaned_data['name']
            selected_date = form.cleaned_data['date']
            selected_year = selected_date.year
            selected_day = selected_date.strftime("%A")
            country = form.cleaned_data['country']  # Fetch the selected country

            year_param = selected_year  # Use the year from the added holiday

            if leave_type == 'country':
                if Holiday.objects.filter(date=selected_date, country=country).exists():
                    messages.warning(request, "⚠️ The date you have entered is already added as a Country Holiday for this country.")
                else:
                    Holiday.objects.create(
                        date=selected_date, 
                        name=name, 
                        day=selected_day, 
                        year=selected_year, 
                        country=country
                    )
                    messages.success(request, "✅ Country holiday added successfully!")

            elif leave_type == 'floating':
                if FloatingHoliday.objects.filter(date=selected_date, country=country).exists():
                    messages.warning(request, "⚠️ The date you have entered is already added as a Floating Holiday for this country.")
                else:
                    FloatingHoliday.objects.create(
                        name=name, 
                        date=selected_date, 
                        year=selected_year, 
                        country=country
                    )
                    messages.success(request, "✅ Floating holiday added successfully!")
            elif leave_type == 'state':
                state = form.cleaned_data['state']
                if StateHoliday.objects.filter(date=selected_date, country=country, state=state).exists():
                    messages.warning(request, "⚠️ This date already exists as a State Holiday for the selected state.")
                else:
                    StateHoliday.objects.create(
                        name=name,
                        date=selected_date,
                        year=selected_year,
                        country=country,
                        state=state
                    )
                    messages.success(request, "✅ State holiday added successfully!")



            # Redirect and preserve filters (send to GET with filters)
            return redirect(f"{reverse('add_holidays')}?country={country.id}&leave_type={leave_type}&year={year_param}")
 
    else:
        # Set initial data for form fields based on filter, so form fields stay sticky
        initial = {}
        if filter_country:
            initial['country'] = filter_country
        if filter_leave_type:
            initial['leave_type'] = filter_leave_type
        form = Holiday_Form(initial=initial)

    # Filter holidays for table by year, country, and type
    filter_kwargs = {'year': filter_year}
    if filter_country:
        filter_kwargs['country__id'] = filter_country

    # For displaying types, filter accordingly:
    if filter_leave_type == 'country':
        country_holidays = Holiday.objects.filter(**filter_kwargs)
        floating_holidays = FloatingHoliday.objects.none()
    elif filter_leave_type == 'floating':
        country_holidays = Holiday.objects.none()
        floating_holidays = FloatingHoliday.objects.filter(**filter_kwargs)
    else:
        country_holidays = Holiday.objects.filter(**filter_kwargs)
        floating_holidays = FloatingHoliday.objects.filter(**filter_kwargs)

    context = {
        'form': form,
        'fixed_holidays': country_holidays,
        'floating_holidays': floating_holidays,
        'current_year': filter_year,
        'filter_country': filter_country,
        'filter_leave_type': filter_leave_type,
        'filter_year': filter_year,
    }
    return render(request, 'add_holidays.html', context)
from django.http import JsonResponse
from .models import Holiday, FloatingHoliday



def filter_holidays_by_year(request, year):
    holiday_type = request.GET.get('type', '')  # Leave type filter ('country', 'floating', or '')
    country_id = request.GET.get('country', '')  # Country id filter as string

    holidays = []

    # Prepare filtering kwargs
    filter_kwargs = {'date__year': year}
    if country_id:
        filter_kwargs['country_id'] = country_id  # filter by country if provided

    if holiday_type == 'country' or holiday_type == '':
        # Get country holidays filtered by year and country (if any)
        country_holidays = Holiday.objects.filter(**filter_kwargs).order_by('date') # Country Holidays are stored in Holiday model
        holidays.extend([{
            'name': holiday.name,
            'date': holiday.date.strftime('%Y-%m-%d'),
            'type':'Country',
            'region': holiday.country.country_name
        } for holiday in country_holidays])

    if holiday_type == 'floating' or holiday_type == '':
        # Get floating holidays filtered by year and country (if any)
        floating_holidays = FloatingHoliday.objects.filter(**filter_kwargs).order_by('date')
        holidays.extend([{
            'name': holiday.name,
            'date': holiday.date.strftime('%Y-%m-%d'),
            'type': 'Floating',
            'region': holiday.country.country_name
        } for holiday in floating_holidays])

    if holiday_type == 'state' or holiday_type == '':
        # Get floating holidays filtered by year and country (if any)
        state_holidays = StateHoliday.objects.filter(**filter_kwargs).order_by('date')
        holidays.extend([{
            'name': holiday.name,
            'date': holiday.date.strftime('%Y-%m-%d'),
            'type': 'State',
            'region': holiday.state.name
        } for holiday in state_holidays])



    return JsonResponse({'holidays': holidays})


from .forms import *
from django.contrib.auth import get_user_model
from django.contrib.auth import update_session_auth_hash

@login_required
def change_password(request):
    """Allow admin to change password without requiring old password."""
    if request.method == 'POST':
        form = AdminPasswordChangeForm(request.POST, user=request.user)  # Pass user
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            user = request.user
            user.set_password(new_password)  # Change password
            user.save()
            update_session_auth_hash(request, user)  # Prevent logout
            return redirect('admin_index')  # Redirect to profile after success
    else:
        form = AdminPasswordChangeForm(user=request.user)  # Pass user to form

    return render(request, 'change_password.html', {'form': form})






def mainbase_view(request):
    return render(request,'admin_partials/admin_base.html')


# from django.http import JsonResponse
# from .models import state  # ✅ Ensure correct import
#
# def get_states_by_country(request):
#     country_id = request.GET.get('country_id')
#     if country_id:
#         states = state.objects.filter(country_id=country_id).values('id', 'name')
#         return JsonResponse({'states': list(states)})
#     return JsonResponse({'states': []})


from django.http import JsonResponse
from django.shortcuts import render
from .models import state  # Ensure the correct model name


def load_states(request):
    country_id = request.GET.get('country_id')
    if country_id:
        states = state.objects.filter(country_id=country_id).order_by('name').values('id', 'name')
        return JsonResponse({'states': list(states)})  # ✅ Return JSON response

    return JsonResponse({'states': []})  # ✅ Return empty list if no country selected

#



from django.shortcuts import render
from django.http import JsonResponse
from .models import LeaveRequest
from django.shortcuts import render
from .models import LeaveRequest, Employees

from django.db.models import Q

from django.db.models import Q
from datetime import date
def get_financial_year_dates(request, country, reference_date=None):
    """
    Calculate financial year start and end based on country reset period.
    Defaults to calendar year if reset period not set.
    """
    if reference_date is None:
        reference_date = date.today()

    current_year = reference_date.year
    try:
        reset_period = HolidayResetPeriod.objects.get(country=country)
        start_month = reset_period.start_month
        start_day = reset_period.start_day

        fy_start_candidate = date(current_year, start_month, start_day)
        if reference_date < fy_start_candidate:
            fy_start = date(current_year - 1, start_month, start_day)
        else:
            fy_start = fy_start_candidate

        try:
            fy_end = fy_start.replace(year=fy_start.year + 1) - timedelta(days=1)
        except ValueError:
            fy_end = fy_start + timedelta(days=365) - timedelta(days=1)

        return fy_start, fy_end

    except HolidayResetPeriod.DoesNotExist:
        return date(current_year, 1, 1), date(current_year, 12, 31)


from django.db.models import Q
from datetime import date
from .models import Country  # Import your Country model

@signin_required
def admin_leave_requests(request):
    countries = Country.objects.all().order_by('country_name')

    country_id_str = request.GET.get('country')
    default_country = countries.filter(country_name='India').first() or (countries.first() if countries else None)

    try:
        country_id = int(country_id_str)
        country_obj = countries.get(id=country_id)
    except (TypeError, ValueError, Country.DoesNotExist):
        country_obj = default_country
        country_id = country_obj.id if country_obj else None

    employee_name_filter = request.GET.get('employee_name', '').strip()
    status_filter = request.GET.get('status', 'Pending').strip()

    year_filter_str = request.GET.get('year', '').strip()

    today = date.today()
    reference_year = None
    try:
        reference_year = int(year_filter_str)
    except Exception:
        reference_year = None

    if reference_year:
        reference_date = date(reference_year, 7, 1)
    else:
        reference_date = today

    # Pass country_obj (Country instance) directly
    fy_start, fy_end = get_financial_year_dates(request, country_obj, reference_date=reference_date)

    leave_requests = LeaveRequest.objects.select_related('employee_user', 'employee_master').filter(
        start_date__lte=fy_end,
        end_date__gte=fy_start,
        employee_master__country=country_obj,
    )

    if employee_name_filter:
        leave_requests = leave_requests.filter(
            Q(employee_master__first_name__icontains=employee_name_filter) |
            Q(employee_master__last_name__icontains=employee_name_filter)
        )

    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    all_leave_requests_country = LeaveRequest.objects.filter(employee_master__country=country_obj)

    available_years = sorted(set(
        all_leave_requests_country.values_list('start_date__year', flat=True)
    ), reverse=True)

    if not reference_year:
        reference_year = fy_start.year

    leave_requests_data = [
        {
            "employee_name": (
                f"{lr.employee_master.first_name} {lr.employee_master.last_name}"
                if lr.employee_master else "Not Available"
            ),
            "employee_email": lr.employee_user.email if lr.employee_user else "Not Available",
            "leave_request": lr,
        }
        for lr in leave_requests.order_by('-start_date')
    ]

    context = {
        "leave_requests_data": leave_requests_data,
        "available_years": available_years,
        "current_year": reference_year,
        "current_employee_name": employee_name_filter,
        "current_status": status_filter,
        "current_country_id": country_id,
        "countries": countries,
        "financial_year_start": fy_start,
        "financial_year_end": fy_end,
    }

    return render(request, "leave_request_display.html", context)
from django.http import JsonResponse
from django.db.models import Q
from .models import LeaveRequest





def filter_leave_requests(request):
    year = request.GET.get('year', None)
    employee_name = request.GET.get('employee_name', '')
    status = request.GET.get('status', '')  # Add this line

    leave_requests = LeaveRequest.objects.select_related('employee_master').all()

    if year:
        leave_requests = leave_requests.filter(start_date__year=year)

    if employee_name:
        leave_requests = leave_requests.filter(
            Q(employee_master__emp_fname__icontains=employee_name) |
            Q(employee_master__emp_lname__icontains=employee_name)
        )

    if status:  # Add this condition
        leave_requests = leave_requests.filter(status=status)

    filtered_data = [
        {
            'id': leave.id,
            'employee_name': f"{leave.employee_master.first_name} {leave.employee_master.last_name}".strip() if leave.employee_master else "Not Available",
            'employee_email': leave.employee_master.personal_email if leave.employee_master else "Not Available",
            'leave_type': leave.leave_type,
            'start_date': leave.start_date.strftime('%Y-%m-%d'),
            'end_date': leave.end_date.strftime('%Y-%m-%d'),
            'reason': leave.reason,
            'status': leave.status,
        }
        for leave in leave_requests
    ]

    return JsonResponse(filtered_data, safe=False)





from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib import messages
import requests

def admin_login(request):
    if request.method == 'POST':
        identifier = request.POST.get('username')  # Username or email
        password = request.POST.get('password')
        recaptcha_response = request.POST.get('g-recaptcha-response')  # Get captcha token from frontend

        #  Verify reCAPTCHA with Google
        data = {
            'secret': settings.RECAPTCHA_PRIVATE_KEY,
            'response': recaptcha_response
        }
        r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=data)
        result = r.json()

        if not result.get('success'):
            messages.error(request, 'Invalid reCAPTCHA. Please try again.')
            return redirect('admin_login')

        #  Continue login if captcha passed
        user = None
        if '@' in identifier: 
            try:
                user = User.objects.get(email=identifier)
                username = user.username
            except User.DoesNotExist:
                messages.error(request, 'Invalid email or username.')
                return redirect('admin_login')
        else:
            username = identifier

        user = authenticate(request, username=username, password=password)

        if user is not None and (user.is_staff or user.is_superuser):
            login(request, user)
            return redirect('admin_index')  # Replace with your actual dashboard route
        else:
            messages.error(request, 'Invalid credentials or insufficient permissions.')

    return render(request, 'admin_login.html')  # Render login page




from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Employees # Import your Employee model

@csrf_exempt


def generate_emp_id(request):
    """
    Generates a unique Employee ID based on the selected employee type code.
    Expects 'employee_type' parameter as the code string, e.g. 'C-', 'I-', 'E-'.
    """
    if request.method == "GET":
        employee_type_code = request.GET.get("employee_type", "").strip()
        print(employee_type_code,"=================================================")
        

        if not employee_type_code:
            return JsonResponse({"error": "Missing employee_type parameter"}, status=400)

        # Filter Employees whose employee_id starts with the employee type code
        last_employee = Employees.objects.filter(employee_id__startswith=employee_type_code).order_by("-employee_id").first()

        if last_employee:
            # Remove prefix code length from employee_id to extract number part
            prefix_len = len(employee_type_code)
            numeric_part = last_employee.employee_id[prefix_len:]
            try:
                last_number = int(numeric_part)
            except ValueError:
                # If parse fails, fallback to default start number
                last_number = 100000
            new_number = last_number + 1
    
        else:
            new_number = 100000  # Start from 100000 if no employees found

        new_employee_id = f"{employee_type_code}{new_number}"
        print(new_employee_id,"=============================================")

        return JsonResponse({"employee_id": new_employee_id}) 

    return JsonResponse({"error": "Invalid request method"}, status=405)

def generate_employee_id(employee_type_code):
    """
    Generates the next unique Employee ID given the employee type code.
    """

     # Change 'app' to your Django app name

    # Clean input
    
    employee_type_code = employee_type_code.strip()
    if not employee_type_code:
        raise ValueError("employee_type_code cannot be empty")

    # Get last employee with this type
    last_employee = Employees.objects.filter(employee_id__startswith=employee_type_code).order_by("-employee_id").first()

    # Decide new number
    if last_employee:
        prefix_len = len(employee_type_code)
        numeric_part = last_employee.employee_id[prefix_len:]
        try:
            last_number = int(numeric_part)
        except (ValueError, TypeError):
            last_number = 100000
        new_number = last_number + 1
    else:
        new_number = 100000

    # Compose new Employee ID
    return f"{employee_type_code}{new_number}"

  # e.g., C-100026

from django.http import JsonResponse
from .models import Holiday, FloatingHoliday

from django.http import JsonResponse
from .models import Holiday, FloatingHoliday

