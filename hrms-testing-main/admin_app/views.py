
from django.shortcuts import render
from django.http import HttpResponse
from.models import *

#
# from django.shortcuts import render
from django.http import HttpResponse
from .models import Role, state, Country ,Employees,Department
from datetime import datetime,date



#
# index page view

def index(request):
    return render(request,'index.html')



def admin_index(request):
    return render(request,'admin_index.html')

# adding employees view

def add_employee(request):
    if request.method == 'POST':
        emp_id = request.POST.get('empid')
        sal_id = request.POST.get('sal')  # Get selected salutation
        f_name = request.POST.get('fname')
        m_name = request.POST.get('mname')
        l_name = request.POST.get('lname')
        emp_valid_from = request.POST.get('empvalidfrom')
        emp_valid_to = request.POST.get('empvalidto')
        emp_c_email = request.POST.get('empcemail')
        emp_p_email = request.POST.get('emppemail')
        emp_mob_ph = request.POST.get('empphone')
        emp_off_ph = request.POST.get('empophone')
        emp_home_ph = request.POST.get('emphphone')
        emp_addr = request.POST.get('empaddr')
        emp_street = request.POST.get('street')
        emp_city = request.POST.get('city')
        pincode=request.POST.get('pincode')
        resume = request.FILES.get('resume')
        certif=request.FILES.get('certif')
        role_id = request.POST.get('role')
        dep = request.POST.get('dep')
        desig=request.POST.get('desig')
        state_id = request.POST.get('state')
        country_id = request.POST.get('country')
        employee_status = request.POST.get('employee_status')
        manager_id = request.POST.get('manager')  # New field for manager
        emp_cp_relation=request.POST.get('empcprelation')
        emp_cp_ph=request.POST.get('empcpph')
        emp_cp_email=request.POST.get('empcpemail')
        emp_cp_name=request.POST.get('empcpname')
        empsalary=request.POST.get('empsalary')

        # # Convert dates
        # emp_valid_from = datetime.strptime(emp_valid_from, '%Y-%m-%d').date()
        # emp_valid_to = datetime.strptime(emp_valid_to, '%Y-%m-%d').date()

        # Set default dates if not provided
        if not emp_valid_from:
            emp_valid_from = date.today()
        else:
            emp_valid_from = datetime.strptime(emp_valid_from, '%Y-%m-%d').date()

        if not emp_valid_to:
            emp_valid_to = date(9999, 12, 31)
        else:
            emp_valid_to = datetime.strptime(emp_valid_to, '%Y-%m-%d').date()

        # Validate inputs
        if not role_id:
            return HttpResponse("Role is required", status=400)
        if emp_valid_from >= emp_valid_to:
            return HttpResponse("Invalid date range", status=400)

        try:
            sal = Salutation.objects.get(id=sal_id)
        except Salutation.DoesNotExist:
            return HttpResponse("Salutation does not exist", status=400)

        try:
            role = Role.objects.get(id=role_id)
        except Role.DoesNotExist:
            return HttpResponse("Role does not exist", status=400)

        try:
            dep = Department.objects.get(id=dep)
        except Department.DoesNotExist:
            return HttpResponse("Department does not exist", status=400)

        try:
            State = state.objects.get(id=state_id)
        except State.DoesNotExist:
            return HttpResponse("State does not exist", status=400)

        try:
            country = Country.objects.get(id=country_id)
        except Country.DoesNotExist:
            return HttpResponse("Country does not exist", status=400)

        # # If manager_id is provided, validate the manager exists
        # if manager_id:
        #
        #
        #     try:
        #         manager = Employees.objects.get(emp_id=manager_id)
        #     except Employees.DoesNotExist:
        #         return HttpResponse("Manager with this ID does not exist", status=400)
        # else:
        #     manager = None  # No manager selected

        # If manager_id is provided, validate the manager exists
        if manager_id:
            if manager_id.lower() == 'null':  # Check for "None" selection
                manager = None
            else:
                try:
                    manager = Employees.objects.get(emp_id=manager_id)
                except Employees.DoesNotExist:
                    return HttpResponse("Manager with this ID does not exist", status=400)
        else:
            manager = None  # No manager selected

        # Save employee
        data = Employees(
            emp_id=emp_id,
            sal=sal,
            emp_fname=f_name,
            emp_mname=m_name,
            emp_lname=l_name,
            emp_val_from=emp_valid_from,
            emp_val_to=emp_valid_to,
            emp_email=emp_c_email,
            emp_pemail=emp_p_email,
            emp_mob_ph=emp_mob_ph,
            emp_off_ph=emp_off_ph,
            emp_home_ph=emp_home_ph,
            emp_addr=emp_addr,
            emp_home_street=emp_street,
            emp_home_city=emp_city,
            pincode=pincode,
            emp_resume=resume,
            role=role,
            dep=dep,
            designation=desig,
            state=State,
            country=country,
            employee_status=employee_status,
            employee_manager=manager, # Assign manager
            emp_certif=certif,
            emp_cp_relation=emp_cp_relation,
            emp_cp_name=emp_cp_name,
            emp_cp_ph=emp_cp_ph,
            emp_cp_email=emp_cp_email,
            emp_base=empsalary,
            # created_by=request.user,  # Set created_by to the current user
            # modified_by=request.user  # Set modified_by to the current user
        )
        data.save()
        # return HttpResponse("Details added successfully.")
        return redirect('add_employee')

    # Fetch roles, departments, states, countries, and employees for the dropdown
    roles = Role.objects.all()
    departments = Department.objects.all()
    states = state.objects.all()
    countries = Country.objects.all()
    salutations = Salutation.objects.all()  # Fetch salutations for the dropdown
    employees = Employees.objects.all()  # All employees to select a manager

    # Provide default dates for the form
    default_valid_from = date.today()
    default_valid_to = date(9999, 12, 31)

    return render(request, 'sample4.html', {
        'roles': roles,
        'departments': departments,
        'states': states,
        'countries': countries,
        'salutations':salutations,  # Pass salutations to the template
        'employees': employees , # Pass employees to the template for selecting manager
        'default_valid_from': default_valid_from,
        'default_valid_to': default_valid_to
    })





# dependent dropdown feature for country and state

from django.http import JsonResponse


# Keep 'state' lowercase, as defined in your model


from .models import state

def get_states(request):
    country_id = request.GET.get('country_id')  # Get country ID from the request
    if country_id:
        states = state.objects.filter(country_id=country_id).values('id', 'name')  # Use 'state' lowercase
        print(states)  # Debugging line
        return JsonResponse(list(states), safe=False)  # Return states as JSON
    return JsonResponse({'error': 'Country not selected'}, status=400)






# # value insertion to Role table

# Role.objects.get_or_create(role_name="Contractor")
# Role.objects.get_or_create(role_name="Employee")
# Role.objects.get_or_create(role_name="Intern")




# # value insertion to salutable table in the model


# Salutation.objects.get_or_create(sal_name="Mr.")
# Salutation.objects.get_or_create(sal_name="Mrs.")


# # value insertion to department table

# Department.objects.get_or_create(dep_name="Software Development")
# Department.objects.get_or_create(dep_name="QA Engineer")
# Department.objects.get_or_create(dep_name="HR")
# Department.objects.get_or_create(dep_name="IT Support")


# #
# #
# # Insert country values to Country table

# country, created = Country.objects.get_or_create(country_name="India")
# #

# # inserting state values to State table in the model


# # List of Indian states
# states = [
#     "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
#     "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand",
#     "Karnataka", "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur",
#     "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
#     "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura",
#     "Uttar Pradesh", "Uttarakhand", "West Bengal", "Andaman and Nicobar Islands",
#     "Chandigarh", "Dadra and Nagar Haveli and Daman and Diu", "Delhi",
#     "Jammu and Kashmir", "Ladakh", "Lakshadweep", "Puducherry"
# ]

# # Insert states
# for state_name in states:
#     state_created = state.objects.get_or_create(
#         name=state_name,
#         country=country
#     )
#     if state_created:
#         print(f"State '{state_name}' created.")
#     else:
#         print(f"State '{state_name}' already exists.")





#
# to download the added employees into a excel sheet we implements the below export_employees_to_excel view

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Employees

def export_employees_to_excel(request):
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Employees'

    # Define the header row
    headers = [
        'Employee ID', 'Salutation', 'First Name', 'Middle Name', 'Last Name',
        'Email', 'Personal Email', 'Mobile Phone', 'Office Phone', 'Home Phone',
        'Valid From', 'Valid To', 'Country', 'State', 'Address', 'Home Street',
        'Home City', 'Pincode', 'Role', 'Department', 'Designation', 'Manager',
        'Employee Status', 'Emergency Contact Name', 'Emergency Contact Phone',
        'Emergency Contact Email', 'Emergency Contact Relation', 'Base Salary',
        'Resume', 'Certificates', 'Created On', 'Modified On', 'Is Deleted',
        'Floating Holidays Balance', 'Floating Holidays Used', 'Total Leaves',
        'Used Leaves'
    ]

    # Write the header row to the sheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Query all employees
    employees = Employees.objects.select_related('sal', 'country', 'state', 'role', 'dep', 'employee_manager')

    # Write employee data to the sheet
    for row_num, employee in enumerate(employees, 2):
        # Convert datetime fields to naive datetime
        created_on = employee.created_on.replace(tzinfo=None) if employee.created_on else ''
        modified_on = employee.modified_on.replace(tzinfo=None) if employee.modified_on else ''

        row = [
            employee.emp_id,
            employee.sal.sal_name if employee.sal else '',
            employee.emp_fname,
            employee.emp_mname,
            employee.emp_lname,
            employee.emp_email,
            employee.emp_pemail,
            employee.emp_mob_ph,
            employee.emp_off_ph,
            employee.emp_home_ph,
            employee.emp_val_from,
            employee.emp_val_to,
            employee.country.country_name if employee.country else '',
            employee.state.name if employee.state else '',
            employee.emp_addr,
            employee.emp_home_street,
            employee.emp_home_city,
            employee.pincode,
            employee.role.role_name if employee.role else '',
            employee.dep.dep_name if employee.dep else '',
            employee.designation,
            f"{employee.employee_manager.emp_fname} {employee.employee_manager.emp_lname}" if employee.employee_manager else '',
            employee.employee_status,
            employee.emp_cp_name,
            employee.emp_cp_ph,
            employee.emp_cp_email,
            employee.emp_cp_relation,
            float(employee.emp_base),
            employee.emp_resume.url if employee.emp_resume else '',
            employee.emp_certif.url if employee.emp_certif else '',
            created_on,
            modified_on,
            employee.is_delete,
            employee.floating_holidays_balance,
            employee.floating_holidays_used,
            employee.emp_total_leaves,
            employee.emp_used_leaves
        ]

        for col_num, cell_value in enumerate(row, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_value

    # Set the response headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response



def list_employees(request):
    #
    # data=Employees.objects.all()
    data = Employees.objects.filter(is_delete=False)  # Only fetch non-deleted employees
    return render(request,'view_employeeslist.html',{'data':data})



#
#
# from .models import Employees, state, Country
# from django.views.generic import UpdateView
# from django.shortcuts import get_object_or_404
# from django.urls import reverse_lazy
# from .models import Employees, Country, state
# from .forms import EmployeeEditForm
#
# class EmployeeUpdateView(UpdateView):
#     model = Employees
#     form_class = EmployeeEditForm
#     template_name = 'employee_update.html'
#     success_url = reverse_lazy('employee_list')
#
#     def form_valid(self, form):
#         response = super().form_valid(form)
#         messages.success(self.request, "Employee details updated successfully.")  # Debugging message
#         return response
#
#     def form_invalid(self, form):
#         messages.error(self.request, "There was an error updating the employee details.")  # Debugging message
#         return super().form_invalid(form)
#
#     def get_object(self, queryset=None):
#         return get_object_or_404(Employees, pk=self.kwargs['pk'])
#
#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#
#         # Populate countries dropdown
#         form.fields['country'].queryset = Country.objects.all()
#
#         # Populate states based on the selected country
#         if self.object and self.object.country:
#             form.fields['state'].queryset = state.objects.filter(country=self.object.country)
#         else:
#             form.fields['state'].queryset = state.objects.none()
#
#         return form




# from django.views.generic import UpdateView
# from django.shortcuts import get_object_or_404
# from django.urls import reverse_lazy
# from django.contrib import messages
# from .models import Employees, Country, state
# from .forms import EmployeeEditForm
#
#
# class EmployeeUpdateView(UpdateView):
#     model = Employees
#     form_class = EmployeeEditForm
#     template_name = 'employee_update.html'
#     success_url = reverse_lazy('employee_list')
#
#     def get_object(self, queryset=None):
#         return get_object_or_404(Employees, pk=self.kwargs['pk'])
#
#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#
#         # Populate countries dropdown
#         form.fields['country'].queryset = Country.objects.all()
#
#         # Populate states based on the selected country
#         if self.object and self.object.country:
#             form.fields['state'].queryset = state.objects.filter(country=self.object.country)
#         else:
#             form.fields['state'].queryset = state.objects.none()
#
#         return form
#
#     def form_valid(self, form):
#         """Handle the form submission"""
#         instance = form.save(commit=False)
#
#         # Handle date fields
#         instance.emp_val_from = self.request.POST.get('emp_val_from')
#         instance.emp_val_to = self.request.POST.get('emp_val_to')
#
#         # Handle file uploads
#         if 'emp_resume' in self.request.FILES:
#             instance.emp_resume = self.request.FILES['emp_resume']
#         if 'emp_certif' in self.request.FILES:
#             instance.emp_certif = self.request.FILES['emp_certif']
#
#         # Handle employee status
#         instance.employee_status = self.request.POST.get('employee_status')
#
#         # Save the instance
#         instance.save()
#
#         messages.success(self.request, 'Employee details updated successfully!')
#         return super().form_valid(form)
#
#     def form_invalid(self, form):
#         """Handle invalid form submission"""
#         messages.error(self.request, 'Please correct the errors below.')
#         return super().form_invalid(form)
#
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['employee'] = self.object  # Add this for accessing current files in template
#         return context


from django.views.generic import UpdateView
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from .models import Employees, Country, state
from .forms import EmployeeEditForm

#
# class EmployeeUpdateView(UpdateView):
#     model = Employees
#     form_class = EmployeeEditForm
#     template_name = 'employee_update.html'
#     success_url = reverse_lazy('employee_list')
#
#     def form_valid(self, form):
#         messages.success(self.request, "Employee details updated successfully.")
#         return super().form_valid(form)
#
#     def form_invalid(self, form):
#         messages.error(self.request, "There was an error updating the employee details.")
#         return super().form_invalid(form)
#
#     def get_queryset(self):
#         return Employees.objects.select_related('country', 'state')
#
#     def get_form(self, form_class=None):
#         form = super().get_form(form_class)
#
#         # Populate countries dropdown
#         form.fields['country'].queryset = Country.objects.all()
#
#         # Populate states dropdown based on selected country
#         if self.object and self.object.country:
#             form.fields['state'].queryset = state.objects.filter(country=self.object.country)
#         else:
#             form.fields['state'].queryset = state.objects.none()
#
#         return form


#
# from django.core.files.storage import default_storage
#
# class EmployeeUpdateView(UpdateView):
#     model = Employees
#     form_class = EmployeeEditForm
#     template_name = 'employee_update.html'
#     success_url = reverse_lazy('employee_list')
#
#     def form_valid(self, form):
#         employee = form.save(commit=False)
#
#         # Handling file upload for emp_resume
#         if self.request.FILES.get('resume'):
#             if employee.emp_resume:
#                 default_storage.delete(employee.emp_resume.path)  # Delete old file
#             employee.emp_resume = self.request.FILES['resume']
#
#         # Handling file upload for emp_certif
#         if self.request.FILES.get('certif'):
#             if employee.emp_certif:
#                 default_storage.delete(employee.emp_certif.path)  # Delete old file
#             employee.emp_certif = self.request.FILES['certif']
#
#         employee.save()
#         messages.success(self.request, "Employee details updated successfully.")
#         return super().form_valid(form)
#
#     def form_invalid(self, form):
#         messages.error(self.request, "There was an error updating the employee details.")
#         return super().form_invalid(form)
#




from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.core.files.storage import default_storage
from .models import Employees
from .forms import EmployeeEditForm

class EmployeeUpdateView(UpdateView):
    model = Employees
    form_class = EmployeeEditForm
    template_name = 'employee_update.html'
    success_url = reverse_lazy('employee_list')

    def form_valid(self, form):
        employee = form.save(commit=False)  # Don't save to DB yet

        # Handle file upload for emp_resume
        resume_file = self.request.FILES.get('emp_resume')
        if resume_file:
            if employee.emp_resume and hasattr(employee.emp_resume, 'path'):
                default_storage.delete(employee.emp_resume.path)  # Delete old file safely
            employee.emp_resume = resume_file

        # Handle file upload for emp_certif
        certif_file = self.request.FILES.get('emp_certif')
        if certif_file:
            if employee.emp_certif and hasattr(employee.emp_certif, 'path'):
                default_storage.delete(employee.emp_certif.path)  # Delete old file safely
            employee.emp_certif = certif_file

        employee.save()  # Save updated employee details
        messages.success(self.request, "Employee details updated successfully.")
        return super().form_valid(form)

    def form_invalid(self, form):
        print("Form Errors:", form.errors.as_json())  # Print detailed errors
        messages.error(self.request, "There was an error updating the employee details.")
        return super().form_invalid(form)






# from django.http import JsonResponse
# from .models import state
#
# def get_states(request):
#     country_id = request.GET.get('country_id')
#     states = State.objects.filter(country_id=country_id).values('id', 'name')
#     return JsonResponse({'states': list(states)})




from django.views.generic import DeleteView
from .models import Employees

class EmployeeDeleteView(DeleteView):
    model = Employees
    template_name = 'delete_employee.html'
    context_object_name = 'data'
    success_url = reverse_lazy('employee_list')  # Redirect after delete

    #soft delete logic mentioned below

    def delete(self, request, *args, **kwargs):
        """Mark the employee as deleted instead of deleting from the database."""
        self.object = self.get_object()
        self.object.delete()  # Call the soft delete method in the model
        return redirect(self.success_url)




from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .models import Employees

def restore_employee(request, pk):
    """Restore a soft-deleted employee."""
    employee = get_object_or_404(Employees, pk=pk)
    if employee.is_delete:
        employee.is_delete = False
        employee.save()
        messages.success(request, f"Employee {employee.emp_fname} {employee.emp_lname} has been restored.")
    else:
        messages.info(request, "This employee is already active.")
    return redirect('employee_list')



#
# above code for listing or displaying only the soft deleted employees

from django.views.generic.list import ListView
from .models import Employees

class DeletedEmployeeListView(ListView):
    model = Employees
    template_name = "deleted_employees_list.html"
    context_object_name = "employees"

    def get_queryset(self):
        return Employees.objects.filter(is_delete=True)





#



from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages

def admin_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')  # Admin username
        password = request.POST.get('password')  # Admin password

        # Authenticate the user using username and password
        user = authenticate(request, username=username, password=password)

        if user is not None and (user.is_staff or user.is_superuser):
            login(request, user)  # Log the admin user in
            return redirect('adminbase')  # Redirect to the admin dashboard
        else:
            messages.error(request, 'Invalid credentials or insufficient permissions')
            return redirect('admin_login')  # Redirect back to login page

    return render(request, 'admin_login.html')  # Render the login page


# Admin Dashboard View (after successful login)
from django.contrib.auth.decorators import login_required

@login_required
def admin_dashboard(request):
    return render(request, 'admin_dashboard.html')  # Render the admin dashboard page


from django.contrib.auth.models import User
from django.http import HttpResponse

def create_admin_user(request):
    # Check if the user is not already created
    if not User.objects.filter(username='admin@example.com').exists():
        admin_user = User.objects.create_superuser('admin@example.com', 'admin@example.com', 'admin_password')
        return HttpResponse('Admin user created successfully!')
    else:
        return HttpResponse('Admin user already exists!')


#

from .models import Holiday
from datetime import date

# # Create or get holidays
# holiday1, created1 = Holiday.objects.get_or_create(
#     date=date(2025, 1, 1),  # New Year's Day
#     name='New Year',
# day='Wednesday'

# )

# holiday2, created2 = Holiday.objects.get_or_create(
#     date=date(2025, 4, 14),  # Vishu
#     name='Vishu',
# day='Monday'

# )

# holiday3, created3 = Holiday.objects.get_or_create(
#     date=date(2025, 4, 18),  # Good Friday
#     name='Good Friday',
# day='Friday'

# )

# holiday4, created4 = Holiday.objects.get_or_create(
#     date=date(2025, 5, 1),  # May Day
#     name='May Day',
# day='Monday'

# )

# holiday5, created5 = Holiday.objects.get_or_create(
#     date=date(2025, 8, 15),  # Independence Day
#     name='Independence Day',
# day='Friday'

# )

# holiday6, created6 = Holiday.objects.get_or_create(
#     date=date(2025, 9, 5),  # Onam
#     name='Onam',
# day='Friday'

# )

# holiday7, created7 = Holiday.objects.get_or_create(
#     date=date(2025, 10, 2),  # Gandhi Jayanthi
#     name='Gandhi Jayanthi',
# day='Thursday'
# )

# holiday8, created8 = Holiday.objects.get_or_create(
#     date=date(2025, 12, 25),  # Christmas
#     name='Christmas',
# day='Thursday'

# )

# # Optionally, print confirmation
# if created1:
#     print("Holiday 'New Year' added successfully.")
# else:
#     print("Holiday 'New Year' already exists.")

# if created2:
#     print("Holiday 'Vishu' added successfully.")
# else:
#     print("Holiday 'Vishu' already exists.")

# if created3:
#     print("Holiday 'Good Friday' added successfully.")
# else:
#     print("Holiday 'Good Friday' already exists.")

# if created4:
#     print("Holiday 'May Day' added successfully.")
# else:
#     print("Holiday 'May Day' already exists.")

# if created5:
#     print("Holiday 'Independence Day' added successfully.")
# else:
#     print("Holiday 'Independence Day' already exists.")

# if created6:
#     print("Holiday 'Onam' added successfully.")
# else:
#     print("Holiday 'Onam' already exists.")

# if created7:
#     print("Holiday 'Gandhi Jayanthi' added successfully.")
# else:
#     print("Holiday 'Gandhi Jayanthi' already exists.")

# if created8:
#     print("Holiday 'Christmas' added successfully.")
# else:
#     print("Holiday 'Christmas' already exists.")



# main leaverequestdisplay view


#
# from django.shortcuts import render
# from .models import LeaveRequest
# from .models import Employees
#
# def leave_request_display(request):
#     # Fetch all leave requests
#     leave_requests = LeaveRequest.objects.all()
#
#     # Prepare data to pass to the template
#     leave_requests_data = []
#     for leave_request in leave_requests:
#         employee_name = "Not Available"
#         employee_email = "Not Available"
#
#         # Check if the leave request is properly linked to an employee
#         if leave_request.employee_master:
#             employee = leave_request.employee_master
#             # Safely access employee fields if the employee exists
#             employee_name = f"{employee.emp_fname} {employee.emp_lname}" if employee.emp_fname and employee.emp_lname else "Name Not Available"
#             employee_email = employee.emp_email if employee.emp_email else "Email Not Available"
#
#         leave_requests_data.append({
#             'leave_request': leave_request,
#             'employee_name': employee_name,
#             'employee_email': employee_email,
#         })
#
#     return render(request, 'leaverequestdisplay.html', {
#         'leave_requests_data': leave_requests_data
#     })
#




#
# from django.shortcuts import render
# from django.db.models.functions import ExtractYear
# from .models import LeaveRequest
#
# def leave_request_display(request):
#     # Extract distinct years from start_date
#     years = LeaveRequest.objects.annotate(year=ExtractYear('start_date')).values_list('year', flat=True).distinct()
#
#     # Fetch all leave requests with related employee data
#     leave_requests = LeaveRequest.objects.select_related('employee_master').all()
#
#     # Prepare data for template
#     leave_requests_data = []
#     for leave_request in leave_requests:
#         employee_name = "Not Available"
#         employee_email = "Not Available"
#
#         # Check if the leave request is properly linked to an employee
#         if leave_request.employee_master:  # ✅ Correct field name
#             employee = leave_request.employee_master
#             # Ensure names and emails are handled properly
#             emp_fname = employee.emp_fname if employee.emp_fname else ""
#             emp_lname = employee.emp_lname if employee.emp_lname else ""
#             employee_name = f"{emp_fname} {emp_lname}".strip() if emp_fname or emp_lname else "Name Not Available"
#             employee_email = employee.emp_email if employee.emp_email else "Email Not Available"
#
#         leave_requests_data.append({
#             'leave_request': leave_request,
#             'employee_name': employee_name,
#             'employee_email': employee_email,
#         })
#
#     return render(request, 'requestdisplay.html', {
#         'leave_requests_data': leave_requests_data,
#         'years': sorted(years, reverse=True),  # Pass available years to the template
#     })
#





# from django.shortcuts import render
# from django.http import JsonResponse
# from django.db.models.functions import ExtractYear
# from django.db.models import Q
# from .models import LeaveRequest
#
# def leave_request_display(request):
#     # Extract distinct years from start_date
#     years = LeaveRequest.objects.annotate(year=ExtractYear('start_date')).values_list('year', flat=True).distinct()
#
#     # Fetch all leave requests
#     leave_requests = LeaveRequest.objects.select_related('employee_master').all()
#
#     leave_requests_data = [
#         {
#             'employee_name': f"{leave.employee_master.emp_fname} {leave.employee_master.emp_lname}".strip() if leave.employee_master else "Not Available",
#             'employee_email': leave.employee_master.emp_email if leave.employee_master else "Not Available",
#             'leave_type': leave.leave_type,
#             'start_date': leave.start_date.strftime('%Y-%m-%d'),
#             'end_date': leave.end_date.strftime('%Y-%m-%d'),
#             'reason': leave.reason,
#             'status': leave.status,
#         }
#         for leave in leave_requests
#     ]
#
#     return render(request, 'requestdisplay.html', {
#         'years': sorted(years, reverse=True),
#         'leave_requests_data': leave_requests_data
#     })
#
#





#
# from django.shortcuts import render
# from django.http import JsonResponse
# from django.db.models.functions import ExtractYear
# from django.db.models import Q
# from .models import LeaveRequest
#
# def leave_request_display(request):
#     # Extract distinct years from start_date
#     years = LeaveRequest.objects.annotate(year=ExtractYear('start_date')).values_list('year', flat=True).distinct()
#
#     # Fetch all leave requests
#     leave_requests = LeaveRequest.objects.select_related('employee_profile').all()
#
#     leave_requests_data = [
#         {
#             'employee_name': f"{leave.employee_profile.emp_fname} {leave.employee_profile.emp_lname}".strip() if leave.employee_profile else "Not Available",
#             'employee_email': leave.employee_profile.emp_email if leave.employee_profile else "Not Available",
#             'leave_type': leave.leave_type,
#             'start_date': leave.start_date.strftime('%Y-%m-%d'),
#             'end_date': leave.end_date.strftime('%Y-%m-%d'),
#             'reason': leave.reason,
#             'status': leave.status,
#         }
#         for leave in leave_requests
#     ]
#
#     return render(request, 'requestdisplay.html', {
#         'years': sorted(years, reverse=True),
#         'leave_requests_data': leave_requests_data
#     })



# leave accept and reject concept(main code)




#
# from django.shortcuts import get_object_or_404, redirect
# from django.http import HttpResponse
# from .models import LeaveRequest
#
# def accept_leave_request(request, leave_request_id):
#     if request.method == "POST":
#         leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
#         leave_request.status = "Approved"
#         leave_request.save()
#         return redirect('leave_request_display')  # Redirect back to the display page
#     return HttpResponse("Invalid Request", status=400)
#
# def reject_leave_request(request, leave_request_id):
#     if request.method == "POST":
#         leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
#         leave_request.status = "Rejected"
#         leave_request.save()
#         return redirect('leave_request_display')  # Redirect back to the display page
#     return HttpResponse("Invalid Request", status=400)
#
#



# accept and reject request with mail sending to the employee view described below

#
# from employee_app.utils import calculate_leave_duration
# from django.core.mail import send_mail
# from django.conf import settings
# from django.shortcuts import get_object_or_404, redirect
# from django.http import HttpResponse
# from .models import LeaveRequest
#
# def accept_leave_request(request, leave_request_id):
#     if request.method == "POST":
#         leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
#         leave_request.status = "Accepted"
#         leave_request.approved_by = request.user  # Add this line to set who approved
#         leave_request.save()
#
#         # Fetch employee's email
#         employee_email = leave_request.employee_user.email
#
#         # Email content
#         subject = "Leave Request Approved"
#         message = (
#             f"Dear {leave_request.employee_user.first_name},\n\n"
#             f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been approved by the admin.\n"
#             f"Status: Approved\n"
#             f"Enjoy your leave!\n\n"
#             f"Best Regards,\nYour Leave Management System"
#         )
#
#         # Send email
#         send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [employee_email])
#
#         return redirect('leave_request_display')  # Redirect back to the display page
#
#     return HttpResponse("Invalid Request", status=400)



from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from datetime import timedelta
from .models import LeaveRequest, Employees
from employee_app.models import Holiday, FloatingHoliday
from employee_app.utils import calculate_leave_duration

def accept_leave_request(request, leave_request_id):
    if request.method == "POST":
        leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
        employee = leave_request.employee_master  # Get the employee instance

        # Calculate leave duration excluding holidays, floating holidays (up to 2), and weekends
        leave_duration = 0
        floating_holidays_used = employee.floating_holidays_used
        max_floating_holidays = 2  # Max floating holidays that can be deducted separately

        current_date = leave_request.start_date
        while current_date <= leave_request.end_date:
            if current_date.weekday() in [5, 6]:  # Skip weekends
                current_date += timedelta(days=1)
                continue
            if current_date in Holiday.objects.values_list('date', flat=True):  # Skip fixed holidays
                current_date += timedelta(days=1)
                continue
            if current_date in FloatingHoliday.objects.values_list('date', flat=True):  # Skip floating holidays
                if floating_holidays_used < max_floating_holidays:
                    floating_holidays_used += 1  # Count floating holiday separately
                    current_date += timedelta(days=1)
                    continue
            leave_duration += 1  # Count as a regular leave day
            current_date += timedelta(days=1)

        # Check if the employee has enough leave balance
        if employee.emp_used_leaves + leave_duration <= 15:
            employee.emp_used_leaves += leave_duration
            employee.floating_holidays_used = floating_holidays_used  # Update floating holidays used
            employee.save()
        else:
            messages.error(request, "Employee cannot exceed the allowed 15 total leave days.")
            return redirect('admin_leave_requests')

        # Update leave request status
        leave_request.status = "Accepted"
        leave_request.approved_by = request.user  # Set who approved it
        leave_request.save()

        # Fetch employee's email
        employee_email = leave_request.employee_user.email

        # Email content
        subject = "Leave Request Approved"
        message = (
            f"Dear {leave_request.employee_user.first_name},\n\n"
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been approved.\n"
            f"Status: Approved\n"
            f"Enjoy your leave!\n\n"
            f"Best Regards,\nYour Leave Management System"
        )

        # Send email notification
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [employee_email])

        return redirect('leave_request_display')

    return HttpResponse("Invalid Request", status=400)


def reject_leave_request(request, leave_request_id):
    if request.method == "POST":
        leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
        leave_request.status = "Rejected"
        leave_request.approved_by = request.user  # Add this line to set who approved
        leave_request.save()

        # Fetch employee's email
        employee_email = leave_request.employee_user.email

        # Email content
        subject = "Leave Request Rejected"
        message = (
            f"Dear {leave_request.employee_user.first_name},\n\n"
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been rejected by the admin.\n"
            f"Status: Rejected\n"
            f"If you have any questions, please contact HR.\n\n"
            f"Best Regards,\nYour Leave Management System"
        )

        # Send email
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [employee_email])

        return redirect('leave_request_display')  # Redirect back to the display page

    return HttpResponse("Invalid Request", status=400)




# def leaverequestdisplay(request):
#     data = LeaveRequest.objects.all()
#     return render(request, 'leaverequestdisplay.html', {'data': data})




# from .models import FloatingHoliday
#
# # List of floating holidays with names, dates, and year
# floating_holidays = [
#
#     {"name": "Maha Shivaratri", "date": "2025-02-26"},
#     {"name": "Holi", "date": "2025-03-14"},
#     {"name": "Ramzan", "date": "2025-03-01"},
#     {"name": "Ram Navami", "date": "2025-04-06"},
#     {"name": "Bakrid", "date": "2025-06-06"}
# ]
#
# # Loop through the list and create or get the objects
# for holiday in floating_holidays:
#     obj, created = FloatingHoliday.objects.get_or_create(
#         name=holiday["name"],
#         date=holiday["date"]
#
#     )
#     if created:
#         print(f"Created new holiday: {holiday['name']}")
#     else:
#         print(f"Holiday already exists: {holiday['name']}")
#


from datetime import date

# floating_holidays = [
#     {"name": "Maha Shivaratri", "date": date(2025, 2, 26)},
#     {"name": "Holi", "date": date(2025, 3, 14)},
#     {"name": "Ramzan", "date": date(2025, 3, 1)},
#     {"name": "Ram Navami", "date": date(2025, 4, 6)},
#     {"name": "Bakrid", "date": date(2025, 6, 6)}
# ]

# # Loop through the list and create or get the objects
# for holiday in floating_holidays:
#     obj, created = FloatingHoliday.objects.get_or_create(
#         name=holiday["name"],
#         date=holiday["date"]
#     )
#     if created:
#         print(f"Created new holiday: {holiday['name']}")
#     else:
#         print(f"Holiday already exists: {holiday['name']}")


from django.shortcuts import redirect
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required

@login_required
def admin_logout(request):
    logout(request)  # Logs out the admin user
    return redirect('admin_login')  # Redirect to the admin login page after logout




from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from .models import Holiday, FloatingHoliday
from .forms import Holiday_Form  # Updated form reference

def add_holidays(request):
    if request.method == "POST":
        form = Holiday_Form(request.POST)  # Updated form reference
        if form.is_valid():
            leave_type = form.cleaned_data['leave_type']
            name = form.cleaned_data['name']
            selected_date = form.cleaned_data['date']  # User-selected date
            selected_year = selected_date.year
            selected_day = selected_date.strftime("%A")  # Get day name (e.g., Monday)

            if leave_type == 'fixed':
                Holiday.objects.create(date=selected_date, name=name, day=selected_day, year=selected_year)
                messages.success(request, "Fixed holiday added successfully!")

            elif leave_type == 'floating':
                FloatingHoliday.objects.create(name=name, date=selected_date, year=selected_year)
                messages.success(request, "Floating holiday added successfully!")

            return redirect('add_holidays')  # Redirect to form after submission

    else:
        form = Holiday_Form()  # Updated form reference

    return render(request, 'add_holidays.html', {'form': form})



from .forms import *
from django.contrib.auth import get_user_model
from django.contrib.auth import update_session_auth_hash

@login_required
def change_password(request):
    """Allow admin to change password without requiring old password."""
    if request.method == 'POST':
        form = AdminPasswordChangeForm(request.POST, user=request.user)  # Pass user
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            user = request.user
            user.set_password(new_password)  # Change password
            user.save()
            update_session_auth_hash(request, user)  # Prevent logout
            return redirect('admin_index')  # Redirect to profile after success
    else:
        form = AdminPasswordChangeForm(user=request.user)  # Pass user to form

    return render(request, 'change_password.html', {'form': form})






def mainbase_view(request):
    return render(request,'admin_partials/admin_base.html')


# from django.http import JsonResponse
# from .models import state  # ✅ Ensure correct import
#
# def get_states_by_country(request):
#     country_id = request.GET.get('country_id')
#     if country_id:
#         states = state.objects.filter(country_id=country_id).values('id', 'name')
#         return JsonResponse({'states': list(states)})
#     return JsonResponse({'states': []})


from django.http import JsonResponse
from django.shortcuts import render
from .models import state  # Ensure the correct model name


def load_states(request):
    country_id = request.GET.get('country_id')
    if country_id:
        states = state.objects.filter(country_id=country_id).order_by('name').values('id', 'name')
        return JsonResponse({'states': list(states)})  # ✅ Return JSON response

    return JsonResponse({'states': []})  # ✅ Return empty list if no country selected

#



from django.shortcuts import render
from django.http import JsonResponse
from .models import LeaveRequest

def admin_leave_requests(request):
    print("🚀 Debug - View is being called!")  # Check if the view is triggered

    leave_requests = LeaveRequest.objects.select_related('employee_user').all()

    # Extract available years for filtering dropdown
    available_years = list(set(leave_requests.values_list('start_date__year', flat=True)))
    print("🚀 Debug - Available Years:", available_years)  # Debugging

    leave_requests_data = [
        {
            "employee_name": (
                f"{leave_request.employee_user.employee_profile.emp_fname} {leave_request.employee_user.employee_profile.emp_lname}"
                if hasattr(leave_request.employee_user, 'employee_profile') else "Not Available"
            ),
            "employee_email": leave_request.employee_user.email,
            "leave_request": leave_request,
        }
        for leave_request in leave_requests
    ]

    return render(request, "leaverequestdisplay.html", {
        "leave_requests_data": leave_requests_data,
        "available_years": sorted(available_years, reverse=True),  # Ensure years are sorted in descending order
    })


from django.http import JsonResponse
from django.db.models import Q
from .models import LeaveRequest


# def filter_leave_requests(request):
#     year = request.GET.get('year', None)
#     employee_name = request.GET.get('employee_name', '')
#
#     leave_requests = LeaveRequest.objects.select_related('employee_master').all()
#
#     if year:
#         leave_requests = leave_requests.filter(start_date__year=year)
#
#     if employee_name:
#         leave_requests = leave_requests.filter(
#             Q(employee_master__emp_fname__icontains=employee_name) |
#             Q(employee_master__emp_lname__icontains=employee_name)
#         )
#
#     filtered_data = [
#         {
#             'id': leave.id,  # ✅ Add this line
#             'employee_name': f"{leave.employee_master.emp_fname} {leave.employee_master.emp_lname}".strip() if leave.employee_master else "Not Available",
#             'employee_email': leave.employee_master.emp_email if leave.employee_master else "Not Available",
#             'leave_type': leave.leave_type,
#             'start_date': leave.start_date.strftime('%Y-%m-%d'),
#             'end_date': leave.end_date.strftime('%Y-%m-%d'),
#             'reason': leave.reason,
#             'status': leave.status,
#         }
#         for leave in leave_requests
#     ]
#
#     return JsonResponse(filtered_data, safe=False)


def filter_leave_requests(request):
    year = request.GET.get('year', None)
    employee_name = request.GET.get('employee_name', '')
    status = request.GET.get('status', '')  # Add this line

    leave_requests = LeaveRequest.objects.select_related('employee_master').all()

    if year:
        leave_requests = leave_requests.filter(start_date__year=year)

    if employee_name:
        leave_requests = leave_requests.filter(
            Q(employee_master__emp_fname__icontains=employee_name) |
            Q(employee_master__emp_lname__icontains=employee_name)
        )

    if status:  # Add this condition
        leave_requests = leave_requests.filter(status=status)

    filtered_data = [
        {
            'id': leave.id,
            'employee_name': f"{leave.employee_master.emp_fname} {leave.employee_master.emp_lname}".strip() if leave.employee_master else "Not Available",
            'employee_email': leave.employee_master.emp_email if leave.employee_master else "Not Available",
            'leave_type': leave.leave_type,
            'start_date': leave.start_date.strftime('%Y-%m-%d'),
            'end_date': leave.end_date.strftime('%Y-%m-%d'),
            'reason': leave.reason,
            'status': leave.status,
        }
        for leave in leave_requests
    ]

    return JsonResponse(filtered_data, safe=False)